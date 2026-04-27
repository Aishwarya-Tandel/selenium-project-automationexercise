import os.path
import openpyxl


class ReadContactUsFormData:
    @staticmethod
    def read_contact_us_form_data():

        path = os.path.join(os.path.dirname(os.path.abspath(__file__)) , "contact_us_form_data.xlsx")

        print(os.path.dirname(os.path.abspath(__file__)))
        print(path)

        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        data_list = []

        for r in range(2 , sheet.max_row +1):
            d_list = []
            for c in range(1 , sheet.max_column + 1):
                data = sheet.cell(r , c).value
                d_list.append(data)

            data_list.append(d_list)

        return data_list


ReadContactUsFormData.read_contact_us_form_data()
